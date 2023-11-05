from curl_cffi.requests import Session  # pip install curl-cffi=0.5.7
from bs4 import BeautifulSoup
from pymongo import MongoClient
import pandas as pd
import numpy as np
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from dotenv import load_dotenv
from datetime import datetime
import time
import random
import os
import json

# FOR DEBUG ONLY - printing unicode chars
import sys
import codecs

sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())
# _________________________________________________________

load_dotenv()  # take environment variables from .env.

MONGODB_URI = os.environ.get("MONGODB_URI")
DB_NAME = os.environ.get("DB_NAME")
COLLECTION_NAME = os.environ.get("COLLECTION_NAME")

BASE_URL = 'https://www.barcodelookup.com/'

# Define the user-agent to impersonate
USER_AGENT_LIST = [
    'chrome99',
    'chrome100',
    'chrome101',
    'chrome104',
    'chrome107',
    'chrome110',
    'chrome99_android',
    'edge99',
    'edge101',
    'safari15_3',
    'safari15_5',
]

CHUNK_SIZE = 5  # 50 100 prod / param of script

INPUT_FILE_PATH = 'input/ean.xlsx'
OUTPUT_FILE_PATH = 'output/asin.xlsx'

# MongoDB setup (.env file later)
client = MongoClient(MONGODB_URI)
db = client[DB_NAME]  # replace with your database name
products_collection = db[COLLECTION_NAME]  # replace with your collection name


def insert_product(product_obj):
    try:
        # Check if jet exists
        existing_product = products_collection.find_one(
            {"_id": product_obj["id"]})
        if not existing_product:
            # In MongoDB, the unique identifier is "_id". We're setting it to the same value as "id".
            product_obj["_id"] = product_obj["id"]
            # Insert the jet object
            products_collection.insert_one({
                **product_obj,
                'createdAt': datetime.now(),
                'updatedAt': datetime.now()
            })
            print(f"-->Added product with id {product_obj['id']}")
        else:
            updated = False  # Flag to check if update occurred
            for key in existing_product:
                if existing_product[key] is None and product_obj.get(key) is not None:
                    products_collection.update_one(
                        {'_id': product_obj["id"]},
                        {
                            '$set': {
                                **product_obj,
                                'updatedAt': datetime.now()
                            }
                        }
                    )
                    print(f"-->Updated product with id {product_obj['id']}")
                    updated = True
                    break  # Exit the loop after the first update
            if not updated:
                print(f"-->Product with id {product_obj['id']} already exists.")
    except Exception as error:
        print("Insertion error:", error)


def get_product_details(html):
    product = {}
    soup = BeautifulSoup(html, 'html.parser')
    prddetails_elem = soup.select_one('div.product-details')
    if not prddetails_elem:
        return None
    product['name'] = prddetails_elem.select_one('h4').text.replace(
        '\n', '') if prddetails_elem else None  # get product name from here
    product['asin'] = None
    product['mpn'] = None
    attributes_elem = soup.select_one('ul#product-attributes')
    if attributes_elem:
        for li in attributes_elem.find_all('li', class_='product-text'):
            text = li.text.strip()
            if "ASIN:" in text:
                asin = text.split("ASIN:")[1].strip()
                product['asin'] = asin
            elif "MPN:" in text:
                mpn = text.split("MPN:")[1].strip()
                product['mpn'] = mpn

    product['stores'] = []
    onlinestores_elem = soup.select_one('div.online-stores')
    # try to get prices by store from here using regX
    store_list = onlinestores_elem.select_one('div.store-list').find('ol')
    for li in store_list.find_all('li'):
        store_name = li.find('span', class_='store-name').text.strip()
        price = li.find('span', class_='store-link').text.strip()
        store = {
            'store': store_name,
            'price': price,
        }
        product['stores'].append(store)
        # from each elem in list get store_name, price & link
    print(f'\t Product details : {product}')
    return product


def process_chunk(chunk, impersonate):
    chunk_data = []
    with Session(impersonate=impersonate) as session:
        print(
            f'---------------------------\nNew session started with {impersonate}\n---------------------------')
        print(chunk.head())
        for ean in chunk['ean']:
            url = f"{BASE_URL}{ean}"
            response = session.get(url)
            print(f'Connecting to {ean}')
            product = get_product_details(response.content)
            if not product:
                continue
            product['id'] = ean
            insert_product(product)
            chunk_data.append(product)
            time.sleep(random.uniform(1, 3))
        session.close()
    return chunk_data


def asin_count():
    # Count documents where 'asin' is not None
    all_products = products_collection.count_documents({})
    asin_products = products_collection.count_documents({'asin': {'$ne': None}})
    print(
        f"\nNumber of documents with 'asin' not None in MongoDb Collection : {asin_products}/{all_products} == {asin_products/all_products}%\n")


def stores_to_string(stores_list):
    return '\n'.join([f"{store['store']} {store['price']}" for store in stores_list])


def export_chunk_to_excel(chunk_data, existing_eans):
    # Filter out products that are already in the Excel file
    new_chunk_data = [product for product in chunk_data if product['id'] not in existing_eans]
    # Transform data for Excel
    new_excel_data = []
    for product in new_chunk_data:
        # Consolidate stores data into a JSON string
        # stores_data = json.dumps(product.get('stores', []), ensure_ascii=False)
        stores_data = stores_to_string(product.get('stores', []))
        # Create a single row for each product
        product_data = {
            'ean': product['id'],
            'name': product.get('name', ''),
            'asin': product.get('asin', ''),
            'stores': stores_data  # Store the JSON string in the 'stores' column
        }
        new_excel_data.append(product_data)
    # Convert to DataFrame
    new_chunk_df = pd.DataFrame(new_excel_data)
    # Append new data to Excel file
    if not new_chunk_df.empty:
        append_to_excel(new_chunk_df)
        # Update the in-memory set of EANs
        existing_eans.update(product['id'] for product in new_chunk_data)
    return existing_eans


def excell_formating():
    workbook = load_workbook(OUTPUT_FILE_PATH)
    worksheet = workbook.active
    # Assuming 'stores' is in column D (you need to adjust this based on your actual column)
    for cell in worksheet['D'][1:]:  # Skip the first row (header)
        cell.alignment = Alignment(wrapText=True)

    # Set the width of the first column
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 50
    worksheet.column_dimensions['C'].width = 17
    worksheet.column_dimensions['D'].width = 30

    # Set a larger row height for all rows except the first one
    for idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):  # Start from the second row
        worksheet.row_dimensions[idx].height = 30

    # # If you want to set the width of all columns, you could loop through them like this:
    # for column_cells in worksheet.columns:
    #     length = max(len(str(cell.value)) for cell in column_cells)
    #     worksheet.column_dimensions[column_cells[0].column_letter].width = length

    workbook.save(OUTPUT_FILE_PATH)

def get_existing_eans():
    if os.path.exists(OUTPUT_FILE_PATH):
        df = pd.read_excel(OUTPUT_FILE_PATH)
        print(f'Output file exists number of existing EANs : {df.shape[0]}')
        return set(df['ean'].astype(str).tolist())
    else:
        print('No output file in specified path a new one will be created once first batch/chunk of Input eans is processed')
        return set()

def append_to_excel(df):
    print(f' ========== Writing {df.shape[0]} products to excell ... ========== ')
    # Append to Excel file if it exists, otherwise create a new one
    if os.path.exists(OUTPUT_FILE_PATH):
        with pd.ExcelWriter(OUTPUT_FILE_PATH, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, index=False, header=False,
                        startrow=writer.sheets['Sheet1'].max_row)
    else:
        with pd.ExcelWriter(OUTPUT_FILE_PATH, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)



if __name__ == '__main__':
    try:
        # Load existing EANs from the output file
        existing_eans = get_existing_eans()

        df = pd.read_excel(INPUT_FILE_PATH)
        total_rows = df.shape[0]
        chunksize = max(1, total_rows // CHUNK_SIZE)
        total_chunks = np.ceil(total_rows / CHUNK_SIZE).astype(int)

        asin_count()

        for i, chunk in enumerate(np.array_split(df, chunksize)):
            # MongoDb DEBUGG : TO CONTINUE COMPLETE RUN ON ALL 14500 EANs
            # if i >= 0 and i <= 72:
            #     continue

            # EXCELL UNIQUE ROWS DEBBUGG : RERUN CODE ON SAME OUTPUT FILE NO DOUBLONS !!
            if i == 15:
                break

            print(f'\n\nProcessing chunk {i+1}/{total_chunks}...')


            chunk_data = []
            
            
            impersonate = random.choice(USER_AGENT_LIST) # [i % len(user_agent_list)]
            chunk_data = process_chunk(chunk, impersonate)
            
            existing_eans = export_chunk_to_excel(chunk_data, existing_eans)
            
            # Format excell document once we are sure it exists
            if i == 0:
                excell_formating()
        excell_formating()
    except Exception as e:
        print("Error:", e)


# # http/socks proxies are supported
# proxies = {"https": "http://localhost:3128"}
# r = requests.get(BASE_URL, impersonate="chrome110", proxies=proxies)
# print(r.content)

# proxies = {"https": "socks://localhost:3128"}
# r = requests.get(BASE_URL, impersonate="chrome110", proxies=proxies)
# print(r.content)
